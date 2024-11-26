import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os


# Funkcja obliczająca progi F
def progiF(n):
    vec = []
    w = 0
    for _ in range(51):
        p = n * pow(10, w)
        if abs(np.floor(p) - n * pow(10, w)) < 1e-7:
            p = np.floor(p)  # Trik numeryczny chroniący przed błędami numerycznymi
        vec.append(p)
        w += 0.1
    return vec


# Funkcja obliczająca progi J
def progiJ():
    vec = []
    w = -8
    for _ in range(51):
        p = pow(10, w)
        if abs(np.floor(p) - pow(10, w)) < 1e-9:
            p = np.floor(p)  # Trik numeryczny chroniący przed błędami numerycznymi
        vec.append(p)
        w += 0.2
    return vec


# Funkcja do zliczania przekroczonych progów J
def count_exceeded_quality_thresholds(value, thresholds):
    counter = 0
    for threshold in reversed(thresholds):  # Iteracja od ostatniego do pierwszego progu
        if value < threshold:
            counter += 1
    return counter


# Funkcja do zapisywania wyników ECDF w Excelu
def save_ecdf_plots_to_excel(input_file, output_file):
    xls = pd.ExcelFile(input_file)
    wb = Workbook()

    # Zakładka 1: Dane surowe
    raw_sheet = wb.active
    raw_sheet.title = "1. dane surowe"

    for sheet_name in xls.sheet_names:
        df = xls.parse(sheet_name)
        for i, col in enumerate(df.columns):
            raw_sheet.cell(row=1, column=i + 1, value=col)
            for j, value in enumerate(df[col]):
                raw_sheet.cell(row=j + 2, column=i + 1, value=value)

    # Pobranie danych do obliczeń ECDF
    real_data = pd.concat([xls.parse("F1_real_best_series"),
                           xls.parse("F2_real_best_series")], axis=1)
    binary_data = pd.concat([xls.parse("F1_binary_best_series"),
                             xls.parse("F2_binary_best_series")], axis=1)

    # Zakładka 2: Progi wykorzystania wywołań F
    progi_f = progiF(1)
    f_sheet = wb.create_sheet(title="2.progi wykorzystania wywołań F")
    f_sheet.append(["Progi wykorzystania budżetu wywołań funkcji celu F (oś odciętych, oś X):"])
    f_sheet.append(["liczba wymiarów przestrzeni poszukiwań: n ="])
    f_sheet.append(["budżet, tj, maksymalna liczba wywołań funkcji celu na jedno uruchomienie"])
    f_sheet.append(["liczba progów: 41; pierwszy próg: n*10**0; ostatni próg: n*10**4"])
    f_sheet.append(["nr progu", "wykładnik potęgi", "wartość progu"])
    for i, p in enumerate(progi_f):
        f_sheet.append([i + 1, np.log10(p), p])

    # Zakładka 3: Progi jakości
    progi_j = progiJ()
    j_sheet = wb.create_sheet(title="3. progi jakości rozwiązania")
    j_sheet.append(["Progi jakości - różnica między wartością optimum a wartością "
                    "najlepszego znalezionego rozwiązania"])
    j_sheet.append(["nr progu", "wartość progu"])
    for i, p in enumerate(progi_j):
        j_sheet.append([i + 1, p])

    # Zakładka 4: Wykres ECDF z danymi
    plot_folder = "ecdf_plots"
    if not os.path.exists(plot_folder):
        os.makedirs(plot_folder)

    # Wykres dla reprezentacji binarnej (progi jakości w przedziale <1; 10>)
    plt.figure(figsize=(10, 6))
    selected_thresholds = [qt for qt in progi_j if 1 <= qt <= 10]
    for qt in selected_thresholds:
        plt.step(binary_data.values.flatten(),
                 np.arange(1, len(binary_data.values.flatten()) + 1) / len(binary_data.values.flatten()),
                 where='post', label=f"{qt:.2f}")
    plt.xlabel("Wartości")
    plt.ylabel("ECDF")
    plt.title("Wykresy ECDF dla reprezentacji binarnej")
    plt.legend()
    plt.grid(True)
    binary_plot_path = os.path.join(plot_folder, "ecdf_binary.png")
    plt.savefig(binary_plot_path)
    plt.close()

    # Wykres dla reprezentacji rzeczywistej (progi jakości w przedziale <1; 10>)
    plt.figure(figsize=(10, 6))
    for qt in selected_thresholds:
        plt.step(real_data.values.flatten(),
                 np.arange(1, len(real_data.values.flatten()) + 1) / len(real_data.values.flatten()),
                 where='post', label=f"{qt:.2f}")
    plt.xlabel("Wartości")
    plt.ylabel("ECDF")
    plt.title("Wykresy ECDF dla reprezentacji rzeczywistej")
    plt.legend()
    plt.grid(True)
    real_plot_path = os.path.join(plot_folder, "ecdf_real.png")
    plt.savefig(real_plot_path)
    plt.close()

    # Dodanie wyników do nowej zakładki "wyniki"
    results_sheet = wb.create_sheet(title="wyniki")

    # Nagłówki - pierwszy wiersz to progi J
    results_sheet.append(["Progi F"] + [f"{qt:.2e}" for qt in sorted(progi_j)])

    # Dane dla reprezentacji binarnej - wyniki dla każdego z progów F
    for pf in progi_f:
        counts = []
        for column in binary_data.columns:
            Fvec = []
            for value in binary_data[column]:
                Fvec.append(count_exceeded_quality_thresholds(value, progi_j))
            # Obliczanie liczby eksperymentów, które przekroczyły każdy z progów jakości
            row_counts = [sum(1 for liczba_progow in Fvec if liczba_progow > i) for i in range(len(progi_j))]
            counts.append(row_counts)
        row = [pf] + [sum(x) for x in zip(*counts)]
        results_sheet.append(row)

    # Pusta linia dla przejrzystości
    results_sheet.append([""])

    # Dane dla reprezentacji rzeczywistej - wyniki dla każdego z progów F
    results_sheet.append(["Progi F"] + [f"{qt:.2e}" for qt in sorted(progi_j)])
    for pf in progi_f:
        counts = []
        for column in real_data.columns:
            Fvec = []
            for value in real_data[column]:
                Fvec.append(count_exceeded_quality_thresholds(value, progi_j))
            # Obliczanie liczby eksperymentów, które przekroczyły każdy z progów jakości
            row_counts = [sum(1 for liczba_progow in Fvec if liczba_progow > i) for i in range(len(progi_j))]
            counts.append(row_counts)
        row = [pf] + [sum(x) for x in zip(*counts)]
        results_sheet.append(row)

    # Dodanie wykresów jako obrazy
    binary_img = Image(binary_plot_path)
    binary_img.width, binary_img.height = 800, 600
    results_sheet.add_image(binary_img, "B6")

    real_img = Image(real_plot_path)
    real_img.width, real_img.height = 800, 600
    results_sheet.add_image(real_img, "O6")

    # Zapisz plik Excel
    wb.save(output_file)


# Główna funkcja
def main():
    input_file = 'results.xlsx'
    output_file = 'ECDF_BBOB_Output.xlsx'

    save_ecdf_plots_to_excel(input_file, output_file)
    print(f"Wyniki ECDF zostały zapisane w pliku {output_file}")


if __name__ == "__main__":
    main()
